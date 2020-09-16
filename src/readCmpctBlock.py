
import openpyxl
import os

# 将backup中的compactBlockValidation文本数据转换为xlsx文件
def ConvertTxt2Xlsx(file_path):
    save_dir = os.getcwd()+"/data/"                 # 获取保存路径
    file_dir, file_full_name = os.path.split(file_path)
    saveName, postName = os.path.splitext(file_full_name)   # 获取文件名
    print("文件目录: {}\n文件全名: {}\n文件名: {}\n后缀名: {}".format(file_dir, file_full_name, saveName, postName ))
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "mysheet"
    row = 2                                                         # 第二行开始写入,第一行留空
    f = open(file_path, 'r')
    line = f.readline()
    while line:
        tmpData = line.split(" ")
        data = []
        for e in tmpData:
            if len(e) >0:
                data.append(e)
        print("当前数据: {}".format(tmpData))
        print("转换后数据: {}".format(data))
        if len(data) >= 9:
            row_data = []
            # row_data.append(data[0])                               # 邻居节点ip和端口号
            row_data.append(data[1])                                 # 区块接收时间
            row_data.append(data[2])                                 # 区块哈希
            row_data.append(data[3])                                 # 区块高度
            row_data.append(data[4])                                 # 本地交易池交易数
            row_data.append(data[5])                                 # 区块交易数
            row_data.append(data[7])                                 # 交易池命中数, 包括data[8]中的数据
            row_data.append(data[8])                                 # orphan_tx 命中数
            if len(data) > 9:
                rs = 1
                if data[9].find('F')!=-1:
                    rs = 0
                row_data.append(rs)                                 # 是否成功在本地重组区块
            print("写入第{} 行数据 : {}".format(row-1, row_data))
            for i in range(len(row_data)):
                worksheet.cell(row, i+1, row_data[i])
            row += 1
        line = f.readline()  
    f.close()
    workbook.save(save_dir+saveName+".xlsx")
    print("数据转换完毕~")

# 将data中所有数据写入xlsx文件中
def writeXlsx(data, dir, title):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "mysheet"
    # 写入title
    row = 1
    for i in len(title):
        worksheet.cell(row, i+1, title[i])
    # 写入数据
    row += 1
    for row_data in data:
        for i in range(len(row_data)):
            worksheet.cell(row, i+1, row_data[i])
        row += 1
    workbook.save(dir)
    print("成功写入文件 {}".format(dir))


# 读取ip地址,时间,哈希值,区块高度,mempool交易量,区块交易量,预填充,本地交易池数量,是否成功重构
def readCompactInfo(dir):
    if os.path.exists(dir) == False:
        print("{} does not exists".format(dir))
        return None
    f = open(dir, 'r')
    line = f.readline()
    data = []
    while line:
        line = line.replace('\n')
        tmpData = line.split(" ")
        for e in tmpData:
            if len(e) >0:
                data.append(e)
        # print("当前数据: {}".format(tmpData))

        print("转换后数据: {}".format(data))
        if len(data) >= 9:
            row_data = []
            row_data.append(data[0])                                      # 邻居节点ip和端口号
            row_data.append(data[1])                                      # 区块接收时间
            row_data.append(data[2])                                      # 区块哈希
            row_data.append((int)(data[3]))                               # 区块高度
            row_data.append((int)(data[4]))                                 # 本地交易池交易数
            row_data.append((int)(data[5]))                                 # 区块交易数
            row_data.append((int)(data[7]))                                 # 本地命中数
            row_data.append((int)(data[8]))                                 # orphan_tx 命中数
            if len(data) > 9:
                row_data.append(data[9]=="SUCCESS")                       # 是否成功在本地重组区块
            else:
                row_data.append((int)(data[5]) ==(int)(data[7])+1 )
            data.append(row_data)
    f.close()
    return data

# 从begin日期到end日期,读取压缩区块的接收信息
def readcmpctBlkInfo(begin, end, dir):
    data = []
    while begin <= end:
        file_dir = dir
        if begin >= 10:
            file_dir += (str)(begin)+"_compactBlockValidation.log"
            tmpres = readCompactInfo(file_dir)
            if tmpres != None:
                data += tmpres
        begin += 1
    return data

def ConvertFromDir(file_dir):
    # root 表示所在目录
    # dirs 表示目录下的所有目录文件
    # files 表示
    for root, dirs, fileNames in os.walk(file_dir):
        for file in fileNames:
            if file.find("compactBlockValidation.log") !=-1:
                file_dir = root + "/"+file
                print("正在转换 {} {}".format(file_dir,dirs))
                ConvertTxt2Xlsx(file_dir)
    print("转换完毕 ")

def exp20200710():
    file_dir = "/home2/zxhu/bitcoin-0.19.0_hzx/experiment20200710/2020-07-"
    save_dir = "/home2/zxhu/bitcoin-0.19.0_hzx/pyScript/data/20200710.xlsx"
    data = readcmpctBlkInfo(13, 24, file_dir)
    title = ["ip-port", "recv_time", "block_hash","block_height", "mempool_tx_cnt", "tx_cnt", "local_hit_cnt", "rebuild_state"]
    writeXlsx(data,save_dir, title)

if __name__ == "__main__":
    exp20200710()